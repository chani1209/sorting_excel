import pandas as pd
from openpyxl import load_workbook
import os
import sys

# 엑셀파일을 읽고, 디렉토리와 제외할 확장자를 입력받아 파일을 삭제하는 클래스
class FileDeleter:
    def __init__(self):
        self.file_name = None
        self.directory_path = None
        self.except_extension_list = []

    def set_file_name(self):
        file_name = input("목록이 있는 파일명을 입력하세요: ")
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        if not os.path.exists(file_name):
            print("파일이 존재하지 않습니다.")
            sys.exit(1)
        self.file_name = file_name

    def set_directory_path(self):
        directory_path = input("경로를 입력하세요: /Users/LGD")
        if not os.path.exists(directory_path):
            print("경로가 존재하지 않습니다.")
            sys.exit(1)
        self.directory_path = directory_path

    def set_except_extension_list(self):
        input_extension_str = input("제외할 확장자를 입력하세요: .png .xlsx .txt 형식으로 소문자로 입력하세요. (공백시 전체) : ")
        self.except_extension_list = self.make_except_extension_list(input_extension_str)

    def make_except_extension_list(self, except_extension):
        except_extension_list = except_extension.split(' ')
        except_extension_list = [x.strip() for x in except_extension_list]
        return except_extension_list

    def is_file_extension_allowed(self, file_path):
        _, file_extension = os.path.splitext(file_path)
        return file_extension.lower() not in self.except_extension_list

    def is_file_name_allowed(self, file_path):
        file_name = os.path.basename(file_path)
        for name in self.allowed_file_name:
            if name in file_name:
                return True
        return False

    def delete_files_with_name_in_directory(self):
        files_to_delete = []
        for root, dirs, files in os.walk(self.directory_path):
            for file in files:
                if file == 'nan':
                    continue
                file_path = os.path.join(root, file)
                if self.is_file_name_allowed(file_path) and self.is_file_extension_allowed(file_path):
                    files_to_delete.append(file_path)

        if not files_to_delete:
            print("삭제할 파일이 없습니다.")
            sys.exit(1)

        print("삭제 대상 파일 리스트:")
        for file_path in files_to_delete:
            print(file_path)

        confirm = input("위의 파일들을 삭제하시겠습니까? (y/n): ")
        if confirm.lower() == 'y':
            for file_path in files_to_delete:
                os.remove(file_path)
                print(f"{file_path}을(를) 삭제했습니다.")
        else:
            print("파일 삭제를 취소했습니다.")

    def run(self):
        try:
            self.set_file_name()
            self.set_directory_path()
            self.set_except_extension_list()

            df = pd.read_excel(self.file_name, header=None)
            self.allowed_file_name = df.iloc[:, 0].tolist()
            self.delete_files_with_name_in_directory()
        except SystemExit:
            pass

# 엑셀파일을 읽고, 파일을 입력받아 그 행을 삭제하는 클래스
class ExcelRowDeleter:
    def __init__(self):
        self.file_name = None
        self.delete_file_name = None
        self.keyword_list = []
        self.file_path_list = []

    def set_file_list_name(self):
        file_name = input("목록이 있는 파일명을 입력하세요: ")
        if not file_name.endswith('.xlsx'):
            file_name += '.xlsx'
        if not os.path.exists(file_name):
            print("파일이 존재하지 않습니다.")
            sys.exit(1)
        self.file_name = file_name

    def set_delete_file_name(self):
        delete_file_name = input("삭제할 파일명을 입력하세요: ")
        if not delete_file_name.endswith('.xlsx'):
            delete_file_name += '.xlsx'
        if not os.path.exists(delete_file_name):
            print("파일이 존재하지 않습니다.")
            sys.exit(1)
        self.delete_file_name = delete_file_name

    def is_title_in_keyword_list(self, title):
        for keyword in self.keyword_list:
            if keyword in title:
                return True
        return False

    def run(self):
        try: 
            self.set_file_list_name()
            self.set_delete_file_name()
            df = pd.read_excel(self.file_name, header=None)
            self.keyword_list = df.iloc[:, 0].tolist()
            workbook = load_workbook(self.delete_file_name)
            sheet = workbook.active
        
            # 삭제할 열을 찾아서 삭제
            columns_to_delete = []
            
            for col_idx, header in enumerate(list(sheet.iter_rows(values_only=True))[0]):
                if header is not None and self.is_title_in_keyword_list(str(header)):
                    columns_to_delete.append(col_idx + 1)
                    print(col_idx + 1)

            if columns_to_delete:
                for col_idx in reversed(columns_to_delete):
                    sheet.delete_cols(col_idx, 1)

                workbook.save(self.delete_file_name)
                print(f"{', '.join(self.keyword_list)}를 포함한 제목을 가진 열을 삭제했습니다.")
            else:
                print(f"{', '.join(self.keyword_list)}를 포함한 제목을 가진 열이 존재하지 않습니다.")

        except SystemExit:
            pass


# 위의 함수들을 하나의 클래스로 묶어서 사용하는 클래스
class ActivateProgram:
    def __init__(self, number) -> None:
        self.program = None
        self.number = number

    def set_program(self, program):
        self.program = program
    
    def run(self):
        self.program.run()


# 메인 함수
if __name__ == '__main__':
    print("프로그램을 선택하세요.")
    print("1. 엑셀파일을 읽고, 디렉토리와 제외할 확장자를 입력받아 파일을 삭제하는 프로그램")
    print("2. 엑셀파일을 읽고, 파일을 입력받아 그 행을 삭제하는 프로그램")
    number = input("번호를 입력하세요: ")
    activate_program = ActivateProgram(number)
    if number == '1':
        activate_program.set_program(FileDeleter())
    elif number == '2':
        activate_program.set_program(ExcelRowDeleter())
    else:
        print("잘못된 입력입니다.")
        exit(1)
    
    activate_program.run()